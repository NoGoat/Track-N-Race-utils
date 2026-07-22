import json
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Hardcoded F1 25 track list with names and circuits
TRACK_NAMES = {
    0:  ('Australian Grand Prix',      'Albert Park Circuit'),
    2:  ('Chinese Grand Prix',         'Shanghai International Circuit'),
    3:  ('Bahrain Grand Prix',         'Bahrain International Circuit'),
    4:  ('Spanish Grand Prix',         'Circuit de Barcelona-Catalunya'),
    5:  ('Monaco Grand Prix',          'Circuit de Monaco'),
    6:  ('Canadian Grand Prix',        'Circuit Gilles Villeneuve'),
    7:  ('British Grand Prix',         'Silverstone Circuit'),
    9:  ('Hungarian Grand Prix',       'Hungaroring'),
    10: ('Belgian Grand Prix',         'Circuit de Spa-Francorchamps'),
    11: ('Italian Grand Prix',         'Autodromo Nazionale Monza'),
    12: ('Singapore Grand Prix',       'Marina Bay Street Circuit'),
    13: ('Japanese Grand Prix',        'Suzuka International Racing Course'),
    14: ('Abu Dhabi Grand Prix',       'Yas Marina Circuit'),
    15: ('United States Grand Prix',   'Circuit of the Americas'),
    16: ('São Paulo Grand Prix',       'Autódromo José Carlos Pace'),
    17: ('Austrian Grand Prix',        'Red Bull Ring'),
    19: ('Mexico City Grand Prix',     'Autódromo Hermanos Rodríguez'),
    20: ('Azerbaijan Grand Prix',      'Baku City Circuit'),
    26: ('Dutch Grand Prix',           'Circuit Zandvoort'),
    27: ('Emilia Romagna Grand Prix',  'Autodromo Enzo e Dino Ferrari'),
    29: ('Saudi Arabian Grand Prix',   'Jeddah Corniche Circuit'),
    30: ('Miami Grand Prix',           'Miami International Autodrome'),
    31: ('Las Vegas Grand Prix',       'Las Vegas Street Circuit'),
    32: ('Qatar Grand Prix',           'Losail International Circuit'),
    39: ('British Grand Prix',         'Silverstone Circuit (Reverse)'),
    40: ('Austrian Grand Prix',        'Red Bull Ring (Reverse)'),
    41: ('Dutch Grand Prix',           'Circuit Zandvoort (Reverse)'),
}

def main():
    print("Scanning telemetry-mapper JSON files...")
    # Find all track json files
    json_files = list(Path(".").glob("track_*.json"))
    print(f"Found {len(json_files)} track JSON files.")

    driven_tracks = {}
    for p in json_files:
        try:
            with open(p, "r", encoding="utf-8") as f:
                data = json.load(f)
            track_id = data.get("track_id")
            if track_id is not None:
                driven_tracks[track_id] = {
                    "session_uid": data.get("session_uid", "—"),
                    "track_length_m": data.get("track_length_m", 0),
                    "points_count": len(data.get("points", [])),
                    "drs_zones_count": len(data.get("drs_events", [])),
                    "speed_traps_count": len(data.get("speed_traps", [])),
                    "sector_crossings_count": len(data.get("sector_crossings", [])),
                    "filename": p.name
                }
        except Exception as e:
            print(f"  Error reading {p.name}: {e}")

    # Build sorted list of all track IDs from the official list plus any found files
    all_track_ids = sorted(list(set(TRACK_NAMES.keys()) | set(driven_tracks.keys())))
    
    print(f"Compiling report for {len(all_track_ids)} unique track IDs...")

    # Create Excel Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "F1 25 Track Progress"

    # Explicitly ensure gridlines are visible
    ws.views.sheetView[0].showGridLines = True

    # Styling colors
    color_title_bg = "1A252C"       # Deep Navy/Charcoal
    color_header_bg = "2C3E50"      # Slate Grey/Navy
    color_accent_red = "FF1801"     # Formula 1 Red
    color_zebra_bg = "F8F9F9"       # Very light grey
    
    # Status styling
    color_done_bg = "E8F8F5"        # Soft Mint Green
    color_done_font = "117A65"      # Dark Mint Green
    color_rem_bg = "FADBD8"         # Soft Rose Red
    color_rem_font = "78281F"       # Dark Rose Red

    # Border definitions
    border_thin_gray = Border(
        left=Side(style='thin', color='E5E7E9'),
        right=Side(style='thin', color='E5E7E9'),
        top=Side(style='thin', color='E5E7E9'),
        bottom=Side(style='thin', color='E5E7E9')
    )

    # 1. Title Block
    ws.row_dimensions[2].height = 25
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 4

    ws["B2"] = "F1 25 — Telemetry & Track Progress Dashboard"
    ws["B2"].font = Font(name="Segoe UI", size=16, bold=True, color="2C3E50")
    
    ws["B3"] = "Analyzes and tracks telemetry-mapped circuits versus outstanding tracks."
    ws["B3"].font = Font(name="Segoe UI", size=10, italic=True, color="7F8C8D")

    # Accent Red Line beneath title (Row 4, Columns B through L)
    for col_idx in range(2, 13):
        col_letter = get_column_letter(col_idx)
        cell = ws[f"{col_letter}4"]
        cell.fill = PatternFill(start_color=color_accent_red, end_color=color_accent_red, fill_type="solid")

    # 2. KPI Summary Cards (Rows 5 to 7)
    ws.row_dimensions[5].height = 16
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 14

    data_start_row = 11
    data_end_row = data_start_row + len(all_track_ids) - 1

    # Card 1: TOTAL TRACKS (B5:C7)
    ws.merge_cells("B5:C5")
    ws["B5"] = "TOTAL TRACKS IN GAME"
    ws["B5"].font = Font(name="Segoe UI", size=8, bold=True, color="5D6D7E")
    ws["B5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B5"].fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")

    ws.merge_cells("B6:C7")
    ws["B6"] = f"=COUNTA(B{data_start_row}:B{data_end_row})"
    ws["B6"].font = Font(name="Segoe UI", size=16, bold=True, color="2C3E50")
    ws["B6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B6"].fill = PatternFill(start_color="F2F4F4", end_color="F2F4F4", fill_type="solid")

    # Card 2: DRIVEN / DONE (D5:E7)
    ws.merge_cells("D5:E5")
    ws["D5"] = "DRIVEN (TELEMETRY DONE)"
    ws["D5"].font = Font(name="Segoe UI", size=8, bold=True, color=color_done_font)
    ws["D5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["D5"].fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")

    ws.merge_cells("D6:E7")
    ws["D6"] = f'=COUNTIF(E{data_start_row}:E{data_end_row}, "Done")'
    ws["D6"].font = Font(name="Segoe UI", size=16, bold=True, color=color_done_font)
    ws["D6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["D6"].fill = PatternFill(start_color="E8F8F5", end_color="E8F8F5", fill_type="solid")

    # Card 3: REMAINING (F5:G7)
    ws.merge_cells("F5:G5")
    ws["F5"] = "REMAINING (PENDING)"
    ws["F5"].font = Font(name="Segoe UI", size=8, bold=True, color=color_rem_font)
    ws["F5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["F5"].fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

    ws.merge_cells("F6:G7")
    ws["F6"] = f'=COUNTIF(E{data_start_row}:E{data_end_row}, "Remaining")'
    ws["F6"].font = Font(name="Segoe UI", size=16, bold=True, color=color_rem_font)
    ws["F6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["F6"].fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

    # Card 4: COMPLETION RATE (H5:I7)
    ws.merge_cells("H5:I5")
    ws["H5"] = "COMPLETION RATE"
    ws["H5"].font = Font(name="Segoe UI", size=8, bold=True, color="2471A3")
    ws["H5"].alignment = Alignment(horizontal="center", vertical="center")
    ws["H5"].fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")

    ws.merge_cells("H6:I7")
    ws["H6"] = f"=D6/B6"
    ws["H6"].font = Font(name="Segoe UI", size=16, bold=True, color="2471A3")
    ws["H6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["H6"].fill = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
    ws["H6"].number_format = "0.0%"

    # Apply borders and background to merged KPI cards cells
    kpi_card_ranges = [("B5:C5", "B6:C7"), ("D5:E5", "D6:E7"), ("F5:G5", "F6:G7"), ("H5:I5", "H6:I7")]
    for header_range, val_range in kpi_card_ranges:
        start_col, start_row = header_range.split(":")[0][0], int(header_range.split(":")[0][1])
        end_col, end_row = val_range.split(":")[1][0], int(val_range.split(":")[1][1:])
        
        start_col_idx = openpyxl.utils.column_index_from_string(start_col)
        end_col_idx = openpyxl.utils.column_index_from_string(end_col)
        
        for r_idx in range(start_row, end_row + 1):
            for c_idx in range(start_col_idx, end_col_idx + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                t = Side(style='thin', color='BDC3C7') if r_idx == start_row else None
                b = Side(style='thin', color='BDC3C7') if r_idx == end_row else None
                l = Side(style='thin', color='BDC3C7') if c_idx == start_col_idx else None
                r = Side(style='thin', color='BDC3C7') if c_idx == end_col_idx else None
                
                cell.border = Border(top=t or cell.border.top, 
                                     bottom=b or cell.border.bottom, 
                                     left=l or cell.border.left, 
                                     right=r or cell.border.right)

    # 3. Table Header (Row 10)
    ws.row_dimensions[10].height = 26
    headers = [
        ("Track ID", "B", "center"),
        ("Grand Prix", "C", "left"),
        ("Circuit Name", "D", "left"),
        ("Status", "E", "center"),
        ("Length (m)", "F", "right"),
        ("Telemetry Points", "G", "right"),
        ("DRS Zones", "H", "center"),
        ("Speed Traps", "I", "center"),
        ("Sector Crossings", "J", "center"),
        ("Session UID", "K", "center"),
        ("Filename", "L", "left")
    ]

    for title, col, align in headers:
        cell = ws[f"{col}10"]
        cell.value = title
        cell.font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=color_header_bg, end_color=color_header_bg, fill_type="solid")
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border = Border(bottom=Side(style='medium', color='1C2833'))

    # 4. Table Data (Row 11 onwards)
    current_row = data_start_row
    for idx, track_id in enumerate(all_track_ids):
        ws.row_dimensions[current_row].height = 20
        is_even = (idx % 2 == 0)
        row_fill = PatternFill(start_color=color_zebra_bg, end_color=color_zebra_bg, fill_type="solid") if is_even else None

        gp_info = TRACK_NAMES.get(track_id, (f"Unknown Grand Prix {track_id}", f"Unknown Circuit {track_id}"))
        gp_name, circuit_name = gp_info

        # Determine details
        if track_id in driven_tracks:
            t_data = driven_tracks[track_id]
            status = "Done"
            length = t_data["track_length_m"]
            points = t_data["points_count"]
            drs = t_data["drs_zones_count"]
            traps = t_data["speed_traps_count"]
            sectors = t_data["sector_crossings_count"]
            session = t_data["session_uid"]
            filename = t_data["filename"]
        else:
            status = "Remaining"
            length = "—"
            points = "—"
            drs = "—"
            traps = "—"
            sectors = "—"
            session = "—"
            filename = "—"

        # B: Track ID
        ws[f"B{current_row}"] = track_id
        ws[f"B{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

        # C: Grand Prix
        ws[f"C{current_row}"] = gp_name
        ws[f"C{current_row}"].alignment = Alignment(horizontal="left", vertical="center")

        # D: Circuit Name
        ws[f"D{current_row}"] = circuit_name
        ws[f"D{current_row}"].alignment = Alignment(horizontal="left", vertical="center")

        # E: Status
        status_cell = ws[f"E{current_row}"]
        status_cell.value = status
        status_cell.alignment = Alignment(horizontal="center", vertical="center")
        if status == "Done":
            status_cell.font = Font(name="Segoe UI", size=10, bold=True, color=color_done_font)
            status_cell.fill = PatternFill(start_color=color_done_bg, end_color=color_done_bg, fill_type="solid")
        else:
            status_cell.font = Font(name="Segoe UI", size=10, bold=True, color=color_rem_font)
            status_cell.fill = PatternFill(start_color=color_rem_bg, end_color=color_rem_bg, fill_type="solid")

        # F: Length (m)
        ws[f"F{current_row}"] = length
        ws[f"F{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        if isinstance(length, int):
            ws[f"F{current_row}"].number_format = "#,##0"

        # G: Telemetry Points
        ws[f"G{current_row}"] = points
        ws[f"G{current_row}"].alignment = Alignment(horizontal="right", vertical="center")
        if isinstance(points, int):
            ws[f"G{current_row}"].number_format = "#,##0"

        # H: DRS Zones
        ws[f"H{current_row}"] = drs
        ws[f"H{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

        # I: Speed Traps
        ws[f"I{current_row}"] = traps
        ws[f"I{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

        # J: Sector Crossings
        ws[f"J{current_row}"] = sectors
        ws[f"J{current_row}"].alignment = Alignment(horizontal="center", vertical="center")

        # K: Session UID
        ws[f"K{current_row}"] = session
        ws[f"K{current_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"K{current_row}"].font = Font(name="Consolas", size=9, color="566573")

        # L: Filename
        ws[f"L{current_row}"] = filename
        ws[f"L{current_row}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"L{current_row}"].font = Font(name="Consolas", size=9, color="566573")

        # Apply Zebra styling, base fonts, and borders to everything except Status
        for title, col, align in headers:
            if col == "E":
                continue # Skip status because it has its own fill and font
            cell = ws[f"{col}{current_row}"]
            cell.border = border_thin_gray
            if col not in ("K", "L"):
                cell.font = Font(name="Segoe UI", size=10, color="2C3E50")
            if row_fill:
                cell.fill = row_fill

        status_cell.border = border_thin_gray

        current_row += 1

    # Double bottom border on the last data row to close the table cleanly
    for title, col, align in headers:
        cell = ws[f"{col}{current_row-1}"]
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=Side(style='double', color='2C3E50')
        )

    # 5. Column Width Auto-Fitting
    print("Auto-fitting column widths...")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 10  # Track ID
    ws.column_dimensions["C"].width = 28  # Grand Prix
    ws.column_dimensions["D"].width = 38  # Circuit Name
    ws.column_dimensions["E"].width = 14  # Status
    ws.column_dimensions["F"].width = 16  # Length
    ws.column_dimensions["G"].width = 18  # Telemetry Points
    ws.column_dimensions["H"].width = 12  # DRS
    ws.column_dimensions["I"].width = 14  # Speed Traps
    ws.column_dimensions["J"].width = 18  # Sectors
    ws.column_dimensions["K"].width = 20  # UID
    ws.column_dimensions["L"].width = 35  # Filename

    # Save to file in workspace root
    out_file = Path("..") / "F1_25_Track_Status.xlsx"
    try:
        wb.save(out_file)
        print(f"Report generated successfully and saved to: {out_file.resolve()}")
    except PermissionError:
        print("\n" + "="*80)
        print("[WARNING] Could not save progress report to F1_25_Track_Status.xlsx!")
        print("The file appears to be open in Microsoft Excel or another program.")
        print("Please close the Excel file and try running the command again.")
        print("="*80 + "\n")


if __name__ == "__main__":
    main()
